import xlrd
import xlwt
from xlutils.copy import copy
from xlwt import Workbook
import xlsxwriter as xw
import openpyxl


def save_xl():
    # 创建 xls 文件对象
    wb = xlwt.Workbook()

    # 创建工作簿
    sh1 = wb.add_sheet('测试表')

    # 然后按照位置来添加数据,第一个参数是行，第二个参数是列
    sh1.write(0, 0, '姓名')
    sh1.write(0, 1, '成绩')
    sh1.write(1, 0, '张三')
    sh1.write(1, 1, 88)
    sh1.write(2, 0, '李四')
    sh1.write(2, 1, 99.5)

    # 最后保存文件即可
    wb.save('../data/test_w.xls')


def load_xl():
    wb = xlrd.open_workbook("../data/test_w.xls")

    # 获取并打印 sheet 数量
    print("sheet 数量:", wb.nsheets)

    # 获取并打印 sheet 名称
    print("sheet 名称:", wb.sheet_names())

    # 根据 sheet 索引获取内容
    sh1 = wb.sheet_by_index(0)
    # sh = wb.sheet_by_name('测试表')

    # 获取并打印该 sheet 行数和列数
    print(u"sheet %s 共 %d 行 %d 列" % (sh1.name, sh1.nrows, sh1.ncols))

    # 获取并打印某个单元格的值
    print("第一行第二列的值为:", sh1.cell_value(0, 1))

    # 获取整行或整列的值
    rows = sh1.row_values(0)  # 获取第一行内容
    cols = sh1.col_values(1)  # 获取第二列内容

    # 打印获取的行列值
    print("第一行的值为:", rows)
    print("第二列的值为:", cols)

    # 获取单元格内容的数据类型
    print("第二行第一列的值类型为:", sh1.cell(1, 0).ctype)

    for r in range(sh1.nrows):
        # 输出指定行
        print(sh1.row(r))


def style_xl():
    wb = xlrd.open_workbook('../data/sty.xls', formatting_info=True)
    new_wb: Workbook = copy(wb)
    sh = new_wb.get_sheet(0)

    style = xlwt.XFStyle()

    font = xlwt.Font()
    font.name = 'Monaco'
    font.bold = True
    font.height = 18 * 20  # 18 号字体
    style.font = font

    borders = xlwt.Borders()
    borders.top = xlwt.Borders.THIN
    borders.bottom = xlwt.Borders.THIN
    borders.left = xlwt.Borders.THIN
    borders.right = xlwt.Borders.THIN
    style.borders = borders

    alig = xlwt.Alignment()
    alig.vert = xlwt.Alignment.VERT_CENTER
    alig.horz = xlwt.Alignment.HORZ_CENTER
    style.alignment = alig

    sh.write(2, 1, 100, style)
    new_wb.save('../data/sty1.xls')


def xlsx_w_xw():
    # 格式可能丢失
    wb = xw.Workbook('../data/test.xlsx')
    sh = wb.add_worksheet('No1')
    # 填充数据
    for i in range(0, 10):
        sh.write(0, i, '列序号 ' + (str(i+1)))
    wb.close()


def xlsx_w_open():
    # 性能不是很稳定
    wb = openpyxl.load_workbook('../data/test.xlsx')
    sh = wb['No1']
    sh['A2'] = 123
    sh['B2'] = 'B2'
    wb.save('../data/test.xlsx')


if __name__ == '__main__':
    # save_xl()
    # load_xl()
    # style_xl()
    # xlsx_w_xw()
    xlsx_w_open()
